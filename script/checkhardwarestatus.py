import paramiko
import openpyxl
import time
import re
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from datetime import datetime
import pytz

# python 버전 3.12.5

# ✅ 장비 정보 직접 정의
devices = [
    {"ip": "111.111.111.111", "username": "admin", "password": "metro"}, #장비 1
    {"ip": "222.222.222.222", "username": "grpvpn", "password": "10fufkckaRo!"}, #장비 2 (장비 1과 이기종)
]

# ✅ Cisco 명령어 실행 함수
def run_commands(ip, username, password, commands):
    results = []
    try:
        client = paramiko.SSHClient()
        client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        client.connect(ip, username=username, password=password, timeout=5)

        shell = client.invoke_shell()
        shell.send("terminal length 0\n")
        time.sleep(1)
        shell.recv(9999)  # 초기 프롬프트 제거

        for cmd in commands:
            shell.send(cmd + '\n')
            time.sleep(2)
            output = shell.recv(99999).decode(errors='ignore')
            # ✅ "show env all" 명령어 자체가 포함된 줄은 모두 제거

            if cmd.lower() == 'show env all' and "Invalid" in output:
                print(f"⚠️ {ip}: 'show env all' 명령어가 실패하여 'show env'로 대체됩니다.")
                cmd = 'show env'
                shell.send(cmd + '\n')
                time.sleep(2)
                output = shell.recv(99999).decode(errors='ignore')

            output_lines = output.splitlines()
            clean_lines = []
            for line in output_lines:
                # 줄에서 공백 제거하고 명령어 포함되면 건너뜀
                if cmd.lower() not in line.strip().lower():
                    clean_lines.append(line)

            clean_output = "\n".join(clean_lines).strip()
            results.append({'command': cmd, 'output': output})

        client.close()
    except Exception as e:
        results.append({'command': 'N/A', 'output': f"Error: {e}"})
    return results


# ✅ 결과 저장


from openpyxl.styles import Alignment

def save_to_excel(results, file_name='result.xlsx'):
    from openpyxl.styles import Alignment, Border, Side
        # 한국 시간
    kr_tz = pytz.timezone('Asia/Seoul')
    now_str = datetime.now(kr_tz).strftime('%Y%m%d_%H%M%S')
    file_name = f"여의도 장비 점검결과_{now_str}.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Command Results"

    # ✅ 헤더 추가
    #ws.append(['IP', 'Device Name', 'Command', 'Output', 'FAN 상태', 'TEMP 상태', 'POWER 상태', '최종판단']) #최종판단은 엑셀 출력에서 제외
    ws.append(['IP', 'Device Name', 'Command', 'Output', 'FAN 상태', 'TEMP 상태', 'POWER 상태'])


    for device in results:
        ip = device['ip']
        # IP나 기타 조건에 맞춰 하드코딩된 장비 이름 할당
        if ip == '111.111.111.111':
            device_name = '장비1'
        elif ip == '222.222.222.222':
            device_name = '장비2'

        for res in device['results']:
            command = res['command']
            output = res['output']

            # 상태 초기값
            fan_status = ''
            temp_status = ''
            power_status = ''
            final_status = ''

            if command.lower() == 'show env all':
                # 개별 항목 판단
                fan_ok = contains_fan_ok(output)
                temp_ok = contains_temp_ok(output)
                power_ok = contains_power_ok(output)

                fan_status = '이상없음' if fan_ok else '이상있음'
                temp_status = '이상없음' if temp_ok else '이상있음'
                power_status = '이상없음' if power_ok else '이상있음'

                # 최종 판단
                final_status = '이상없음' if fan_ok and temp_ok and power_ok else '이상있음'

            
            if command.lower() == 'show env':
                # 개별 항목 판단
                fan_ok = parse_fan_status(output)
                power_ok = parse_power_status(output)
                temp_ok = parse_temp_status(output)
    
                 # 상태 표시
                fan_status = '이상없음' if fan_ok else '이상있음'
                power_status = '이상없음' if power_ok else '이상있음'
                temp_status = '이상없음' if temp_ok else '이상있음'
    
                 # 최종 판단
                final_status = '이상없음' if fan_ok and power_ok and temp_ok else '이상있음'


            row = [ip,device_name, command, output, fan_status, temp_status, power_status] #최종판단은 엑셀출력에서 제외
            ws.append(row)

            # 셀 정렬 설정
            row_idx = ws.max_row

            for col in range(1, 8):
                cell = ws.cell(row=row_idx, column=col)
                if col == 4:  # Output
                    cell.alignment = Alignment(wrap_text=True, vertical="center")
                else:
        # 나머지 열들은 가로, 세로 중앙 정렬
                    cell.alignment = Alignment(horizontal="center", vertical="center")

    # 컬럼 너비 자동 조정
    for i, column_cells in enumerate(ws.columns, 1):
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        col_letter = openpyxl.utils.get_column_letter(i)
        ws.column_dimensions[col_letter].width = min(max_length + 5, 80)

    wb.save(file_name)
    print(f"✅ 엑셀 저장 완료: {file_name}")


def find_line(text, keyword):
    for line in text.splitlines():
        if keyword.lower() in line.lower():
            return line.lower()
    return ''

def parse_fan_status(output):
    # 팬 상태 파싱 (예: Cooling Fan 1: 4864 RPM - OK)
    fan_pattern = re.compile(r'Cooling Fan \d+: (\d+) RPM - (OK|NOT OK)')
    fans = fan_pattern.findall(output)
    
    fan_ok = True
    for fan, status in fans:
        # 상태가 "NOT OK"이면 팬이 이상으로 판단
        if status.lower() != "ok":
            fan_ok = False
            break
    
    return fan_ok

def parse_power_status(output):
    # 전원 상태 파싱 (예: Left Slot (PS0): Present / OK)
    power_pattern = re.compile(r'(Left Slot|Right Slot) \((PS0|PS1)\): (Present|OK)')
    power_supplies = power_pattern.findall(output)
    
    power_ok = True
    for slot, ps, status in power_supplies:
        # 상태가 "Present" 또는 "OK"가 아닌 경우 문제로 판단
        if status.lower() not in ["present", "ok"]:
            power_ok = False
            break
    
    return power_ok

def parse_temp_status(output):
    # 온도 상태 파싱 (온도 정보가 포함되어 있을 경우)
    # 예시로 임의의 온도 정보를 기준으로 "OK" 판단을 합니다.
    # 실제로는 show env 명령어에 온도 정보가 포함되면 여기에 추가적인 규칙을 만들 수 있습니다.
    temp_pattern = re.compile(r'Temperature: (\d+)C - (OK|NOT OK)')
    temps = temp_pattern.findall(output)
    
    temp_ok = True
    for temp, status in temps:
        # 상태가 "NOT OK"이면 온도가 이상으로 판단
        if status.lower() != "ok":
            temp_ok = False
            break
    
    return temp_ok



def contains_fan_ok(output):
    return contains_ok_line(output, 'fan')

def contains_temp_ok(output):
    return contains_ok_line(output, 'temperature')

def contains_power_ok(output):
    output_lower = output.lower()
    lines = output_lower.strip().splitlines()

    # 1️⃣ 문장형 검사 (power is ok 등)
    for line in lines:
        if 'power' in line and 'ok' in line:
            return True

    # 2️⃣ 테이블형 검사
    # 헤더 줄 탐색
    header_line = None
    data_lines = []
    for i, line in enumerate(lines):
        if 'sys pwr' in line and 'poe pwr' in line:
            header_line = line
            # 데이터는 보통 헤더 바로 밑줄부터 시작하니까 그 다음 두 줄 검사
            data_lines = lines[i+2:]
            break

    if not header_line:
        return False  # 테이블 형식도 없으면 이상

    # 헤더 열 위치 파악
    headers = re.split(r'\s{2,}', header_line.strip())
    try:
        sys_idx = headers.index('sys pwr')
        poe_idx = headers.index('poe pwr')
    except ValueError:
        return False  # 열 이름 못 찾으면 이상

    # 데이터 줄 파싱
    for line in data_lines:
        cols = re.split(r'\s{2,}', line.strip())
        if len(cols) <= max(sys_idx, poe_idx):
            continue

        sys_status = cols[sys_idx].strip()
        poe_status = cols[poe_idx].strip()

        if sys_status.lower() == 'good' and poe_status.lower() == 'good':
            return True

    return False
def contains_ok_line(text, keyword):
    for line in text.splitlines():
        if keyword.lower() in line.lower() and 'ok' in line.lower():
            return True
    return False



        

# ✅ 메인 실행
def main():
    results = []
    commands = [
        "show env all",  # ✅ 이 명령어만 남김
    ]

    for device in devices:
        print(f"🔌 접속 중: {device['ip']}")
        command_results = run_commands(device['ip'], device['username'], device['password'], commands)
        results.append({'ip': device['ip'], 'results': command_results})

    save_to_excel(results)
    print("✅ 점검 완료! result.xlsx에 저장되었습니다.")


if __name__ == '__main__':
    main()


